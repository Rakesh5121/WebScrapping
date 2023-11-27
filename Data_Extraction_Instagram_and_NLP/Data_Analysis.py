
import os
import openpyxl

cur_dir = os.getcwd()
MD_dir = cur_dir + r"\MasterDictionary\\"
SW_dir = cur_dir + r"\StopWords\\"
TF_dir = cur_dir + r"\Text_File\\"

SW_files = [x for x in os.listdir(SW_dir) if x.endswith(".txt")]
TF_files = [x for x in os.listdir(TF_dir) if x.endswith(".txt")]
Pronoun = ["i", 'we', "my", "ours", "us"]
Vowels = ["a", "e", "i", "o", "u"]
SW = []
PW = []
NW = []

dataframe = openpyxl.load_workbook("Output_Data_Structure.xlsx")
df = dataframe.active

def StopWord():
    for x in SW_files:
        with open(SW_dir + x, "r+") as f:
            y = f.readlines()
            for line in y:
                SW.append(line.replace("\n", ""))
    print(len(SW))
    for x in SW:
        SW.remove(x)
        SW.append(x.lower())
    print(len(SW))


def Positive_AND_Negative_WORD():
    with open(MD_dir + "positive-words.txt", "r+") as f:
        y = f.readlines()
        for line in y:
            PW.append(line.replace("\n", ""))
    print(len(PW))
    for x in PW:
        PW.remove(x)
        PW.append(x.lower())
    print(len(PW))
    with open(MD_dir + "negative-words.txt", "r+") as f:
        y = f.readlines()
        for line in y:
            NW.append(line.replace("\n", ""))
    print(len(NW))
    for x in NW:
        NW.remove(x)
        NW.append(x.lower())
    print(len(NW))


def Sentimental_Analysis(file_path=None):
    if file_path:
        Pos_count = 0
        Neg_count = 0
        Total_word = 0
        with open(TF_dir + file_path, "r+", encoding="utf-8") as f:
            lines = f.readlines()
            for line in lines:
                words = line.split(" ")
                for word in words:
                    if word.lower() not in SW:
                        Total_word += 1
                        if word.lower() in PW:
                            Pos_count += 1
                        if word.lower() in NW:
                            Neg_count += 1
        print("********************************************************")
        print(f"*****     FILE: {file_path}      *****")
        print("********************************************************")
        print(f"Positive Score: {Pos_count}   Negative Score: {Neg_count}  Total Word: {Total_word}")
        Pol = ((Pos_count - Neg_count) / (Pos_count + Neg_count + 0.000001))
        print(f'Polarity : {Pol}')
        Sub = ((Pos_count + Neg_count) / (Total_word + 0.000001))
        print(f'Subjectivity Score : {Sub}')
        return Pos_count, Neg_count, Pol, Sub
    else:
        print("Provide the file path")


def Average_word_per_sentance_and_average_character_per_word(file_path=None):
    total_sentance = 0
    total_word = 0
    total_character = 0
    pronoun_count = 0
    sylleble_count = 0
    complex_word_count = 0
    text = ""
    if file_path:
        with open(TF_dir + file_path, "r+", encoding="utf-8") as f:
            lines = f.readlines()
            for line in lines:
                text += line.replace("\n", " ")
            sentances = text.replace("?", ".").split(".")
            total_sentance = len(sentances)
            words = " ".join(sentances).split(" ")
            for w in words:
                wi = w.replace("!", "").replace(",", "")
                if wi.lower() in Pronoun and wi!= "US":
                    #print(wi)
                    pronoun_count += 1
                count = 0
                for x in wi.lower():
                    if x in Vowels and not (x.endswith("es") or x.endswith("ed")):
                        sylleble_count += 1
                        count += 1
                if count>2:
                    complex_word_count += 1
                words.remove(w)
                if wi not in SW and len(wi):
                    words.append(wi)
            total_word = len(words)
            for x in words:
                total_character += len(x)
        print("*************************************************************************************************")
        print(f'Total Sentance: {total_sentance}  Total word:  {total_word}  Total Character: {total_character}')
        AWS = total_word / total_sentance
        ACW = total_character / total_word
        CP = complex_word_count/ total_word
        Fog_Index = 0.4 * (AWS +CP)
        SPW = sylleble_count/total_word
        print(f'Average sentence length:  {AWS}')
        print(f'Average word length:  {ACW}')
        print(f'Pronoun Count: {pronoun_count}')
        print(f'Syllebale Count: {sylleble_count}')
        print(f"Complex Word Count: {complex_word_count}")
        print(f"Fog Index: {Fog_Index}")
        print("*************************************************************************************************")
        print("\n")
        return AWS, CP, Fog_Index, complex_word_count, total_word, SPW, pronoun_count, ACW
    else:
        print("provide the file path")


StopWord()
Positive_AND_Negative_WORD()

for i in range(2, df.max_row + 1):
    url_id = df.cell(row=i, column=1)
    file = str(int(url_id.value)) + ".txt"

    if  file in TF_files:
        Pos_count, Neg_count, Pol, Sub = Sentimental_Analysis(file)
        AWS, CP, Fog_Index, complex_word_count, total_word, SPW, pronoun_count, ACW = Average_word_per_sentance_and_average_character_per_word(file)
        df.cell(row=i, column=3).value = Pos_count
        df.cell(row=i, column=4).value = -Neg_count
        df.cell(row=i, column=5).value = Pol
        df.cell(row=i, column=6).value = Sub
        df.cell(row=i, column=7).value = AWS
        df.cell(row=i, column=8).value = CP
        df.cell(row=i, column=9).value = Fog_Index
        df.cell(row=i, column=10).value = AWS
        df.cell(row=i, column=11).value = complex_word_count
        df.cell(row=i, column=12).value = total_word
        df.cell(row=i, column=13).value = SPW
        df.cell(row=i, column=14).value = pronoun_count
        df.cell(row=i, column=15).value = ACW

    else:
        print("Http was not available!!!")
        for j in range(3,16):
            df.cell(row=i, column=j).value = 0
        df.cell(row=i, column=16).value = "Page Not Found"
dataframe.save("Output_Data_Structure.xlsx")

