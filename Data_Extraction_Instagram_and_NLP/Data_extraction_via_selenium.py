import os

from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

# Setup chrome driver
driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("Input.xlsx")
dir_path = r"C:\Users\423rk\Downloads\Python\Coffer_Assignment\Data_Extraction_and_NLP\Text_File"
# Define variable to read sheet
dataframe1 = dataframe.active
# Iterate the loop to read the cell values
wait = WebDriverWait(driver, 60)
wrong_url = []
for i in range(2, dataframe1.max_row + 1):
    print("Iteration: " + str(i))
    url_id = dataframe1.cell(row=i, column=1)
    url = dataframe1.cell(row=i, column=2)
    print(str(int(url_id.value)) + " : " + str(url.value))
    # Navigate to the url
    driver.get(str(url.value))
    title = driver.title
    # Find all div elements with specific class name
    div_elements = driver.find_elements(By.CLASS_NAME, "td-post-content")
    #Extracting only paragraph

    #div_elements = wait.until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, "div.td-post-content p")))
    if not "Page not found - Blackcoffer Insights" in title:
        with open(dir_path + "\\" + str(int(url_id.value))+".txt", "w+", encoding="utf-8") as f:
            print("Title:" + str(title))
            f.write(str(title) + "\n")
            # Iterate over the divs
            for div_element in div_elements:
                print(div_element.tag_name)
                f.write("\n")
                f.write(div_element.text)
                #print(div_element.text)
    else:
        wrong_url.append(url_id.value)
# Close the driver
driver.quit()
for x in wrong_url:
    os.remove(dir_path + "\\" + str(int(x))+".txt")